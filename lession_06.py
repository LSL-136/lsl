# -*- coding:utf-8 -*-
import openpyxl
import requests
#
def read_data(filename, sheetname):
    wb = openpyxl.load_workbook(filename)  # 获取Excel表格对象
    sheet = wb[sheetname]  # 获取表单对象
    case_list = []
    max_row = sheet.max_row  # 获取测试用例的最大行号
    for i in range(2, max_row+1):
        case_dict = dict(
        case_id=sheet.cell(row=i, column=1).value,  # 行列找对应的单元格，通过.value获取单元格的值
        url=sheet.cell(row=i, column=5).value,
        data=sheet.cell(row=i, column=6).value,
        expected=sheet.cell(row=i, column=7).value
        )  # 使用字典封装
        case_list.append(case_dict)
    return case_list  # 使用返回值获取测试用例
cases = read_data('test_case_api.xlsx', 'login')  # 定义一个变量获取


# 发送接口请求
def api_request(api_url, api_data):  # 定义一个接口函数
    qcd_headers_register = {"X-Lemonban-Media-Type": "lemonban.v2", "Content-Type": "application/json"}
    response = requests.post(url=api_url, json=api_data, headers=qcd_headers_register)  # 给请求的结果赋值
    return response.json()  # 接口函数的返回值

# 对数据重新写入
def write_data(filename, sheetname, row, column, final_result):
    wb = openpyxl.load_workbook(filename)  # 获取工作薄
    sheet = wb[sheetname]  # 获取表单
    sheet.cell(row=row, column=column).value = final_result  # 获数据重写
    wb.save(filename)  # 保存数据，确保工作薄是不打开的状态下才可以执行成功

def execute_result(filename,sheetname):
    cases = read_data(filename,sheetname)  # 变量接受返回值
    for case in cases:  # 遍历cases并获取cases值
        case_id = case.get("case_id")  # 字典取值
        url = case.get("url")
        data = case['data']
        # print(type(data))
        data = eval(data)  # 脱掉引号的外壳
        # print(type(data))
        expected = case.get("expected")
        # print(type(expected))
        expected = eval(expected)
        # print(type(expected))
        real_result = api_request(api_url=url, api_data=data)  # 把用例的url，data传给接口api_request,真实结果
        real_msg = real_result['msg']  # 获取真实的msg结果
        expected_msg = expected.get("msg")  # 获取期望的msg结果
        print("真实执行结果:{}".format(real_msg))
        print("预期执行结果:{}".format(expected_msg))
        if real_msg == expected_msg:  # 判断真实结果和预期结果对比
            print("这{}条测试用例执行通过!".format(case_id))
            final_result = "Passed"  # 回写数据保存到表格中
        else:
            print("这{}条测试用例不通过！".format(case_id))
            final_result = "Failed"
        print("*" * 20)
        write_data(filename, sheetname, case_id+1, 8, final_result)
execute_result("test_case_api.xlsx",'login')


